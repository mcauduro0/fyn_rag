"""
XLSX Document Processor.
Extracts data from Microsoft Excel spreadsheets.
"""

import logging
from typing import Dict, Any, List
from openpyxl import load_workbook

from app.data.processors.base_processor import BaseProcessor

logger = logging.getLogger(__name__)


class XLSXProcessor(BaseProcessor):
    """Processor for XLSX documents."""
    
    def process(self, file_path: str, max_rows_per_sheet: int = 1000) -> Dict[str, Any]:
        """
        Extract data from an XLSX file.
        
        Args:
            file_path: Path to the XLSX file
            max_rows_per_sheet: Maximum rows to extract per sheet
            
        Returns:
            Dictionary containing extracted data and metadata
        """
        try:
            # Validate file
            path = self._validate_file(file_path, ['.xlsx', '.XLSX'])
            
            logger.info(f"Processing XLSX: {path}")
            
            # Load workbook
            wb = load_workbook(str(path), data_only=True, read_only=True)
            
            # Extract data from all sheets
            sheets_data = []
            total_rows = 0
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Get sheet data
                rows = []
                for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if idx > max_rows_per_sheet:
                        logger.warning(f"Sheet '{sheet_name}' exceeded max rows ({max_rows_per_sheet}), truncating")
                        break
                    
                    # Convert row to strings and filter out empty rows
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_data):  # Skip completely empty rows
                        rows.append(row_data)
                
                if rows:
                    sheets_data.append({
                        "sheet_name": sheet_name,
                        "num_rows": len(rows),
                        "num_columns": len(rows[0]) if rows else 0,
                        "data": rows
                    })
                    total_rows += len(rows)
            
            # Convert to text format
            text_content = []
            for sheet_info in sheets_data:
                text_content.append(f"=== Sheet: {sheet_info['sheet_name']} ===")
                text_content.append(f"Rows: {sheet_info['num_rows']}, Columns: {sheet_info['num_columns']}")
                text_content.append("")
                
                # Add rows as text
                for row in sheet_info['data']:
                    text_content.append(" | ".join(row))
                
                text_content.append("")
            
            full_text = "\n".join(text_content)
            
            # Extract metadata
            metadata = {
                "num_sheets": len(sheets_data),
                "total_rows": total_rows,
                "file_name": path.name,
                "file_size_bytes": path.stat().st_size,
                "sheets": [
                    {
                        "name": s["sheet_name"],
                        "rows": s["num_rows"],
                        "columns": s["num_columns"]
                    }
                    for s in sheets_data
                ]
            }
            
            logger.info(f"Successfully extracted {total_rows} rows from {len(sheets_data)} sheets")
            
            wb.close()
            
            return self._success_response(full_text, metadata, path)
            
        except Exception as e:
            return self._error_response(e, file_path)


if __name__ == "__main__":
    # Test the XLSX processor
    logging.basicConfig(level=logging.INFO)
    
    processor = XLSXProcessor()
    
    # Test with a sample XLSX (you'll need to provide a path)
    test_file = "sample.xlsx"
    
    result = processor.process(test_file)
    
    if result["success"]:
        print(f"\nSuccessfully processed XLSX:")
        print(f"Sheets: {result['metadata']['num_sheets']}")
        print(f"Total rows: {result['metadata']['total_rows']}")
        print(f"Content length: {result['content_length']} characters")
        print(f"\nSheet details:")
        for sheet in result['metadata']['sheets']:
            print(f"  - {sheet['name']}: {sheet['rows']} rows x {sheet['columns']} columns")
    else:
        print(f"\nError: {result['error']}")
